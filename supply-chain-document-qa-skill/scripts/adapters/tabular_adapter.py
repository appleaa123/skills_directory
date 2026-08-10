#!/usr/bin/env python3
"""
Tabular adapter: ingests CSV, JSON (a list of objects, or {"records": [...]})
and .xlsx files -- one record per row/object. This is the path for database
exports (orders, shipments, supplier contacts) and for anything a data
pipeline or MCP-connected tool has already dumped to a file: once it's a file
on disk in one of these formats, this skill treats it identically regardless
of where it came from.
"""
import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from records import make_record

SOURCE_TYPE = "tabular"

# Column names (case-insensitive) checked, in order, for a per-row date value.
_DATE_COLUMN_CANDIDATES = ("date", "order_date", "shipment_date", "due_date", "created_at", "timestamp")


def parse(path: Path) -> list:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv(path)
    elif suffix == ".json":
        rows = _read_json(path)
    elif suffix == ".xlsx":
        rows = _read_xlsx(path)
    else:
        raise ValueError(f"tabular_adapter cannot handle extension {suffix!r}")

    records = []
    for row_index, row in enumerate(rows):
        text = _row_to_text(row)
        if not text.strip():
            continue
        date = _find_date(row)
        title = f"{path.stem} row {row_index + 1}"
        records.append(make_record(
            source_path=str(path), source_type=SOURCE_TYPE, text=text,
            title=title, date=date,
            metadata={"row_index": row_index, **{str(k): v for k, v in row.items()}},
            doc_id=f"{SOURCE_TYPE}:{path.name}:{row_index}",
        ))
    return records


def _read_csv(path: Path) -> list:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _read_json(path: Path) -> list:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict) and "records" in data:
        data = data["records"]
    if not isinstance(data, list):
        raise ValueError(f"{path}: expected a JSON list or {{'records': [...]}}, got {type(data).__name__}")
    return [row if isinstance(row, dict) else {"value": row} for row in data]


def _read_xlsx(path: Path) -> list:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        sheet = wb.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(next(rows_iter))]
        except StopIteration:
            return []
        rows = []
        for values in rows_iter:
            if all(v is None for v in values):
                continue
            rows.append({header[i]: values[i] for i in range(min(len(header), len(values)))})
        return rows
    finally:
        wb.close()


def _row_to_text(row: dict) -> str:
    return "\n".join(f"{key}: {value}" for key, value in row.items() if value not in (None, ""))


def _find_date(row: dict) -> str:
    lower_keys = {k.lower(): k for k in row.keys()}
    for candidate in _DATE_COLUMN_CANDIDATES:
        if candidate in lower_keys:
            value = row[lower_keys[candidate]]
            return str(value) if value is not None else ""
    return ""
